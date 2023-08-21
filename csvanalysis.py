from openpyxl import load_workbook
from matplotlib import pyplot as plt

#column_one A
wb = load_workbook("Book1.xlsx") #workbookname
ws = wb['Sheet6'] #worksheetname
column_one = ws['A'] #ColumnA
column_one_list = [column_one[x].value for x in range(len(column_one))] #column_one A as a list
print(column_one_list)

#column_two B
column_two=ws['B']#columnB
column_two_list=[column_two[x].value for x in range(len(column_two))]#column_two B as a list
print(column_two_list)
print('There are',end=' ')
print(len(column_two_list),end=' ') 
print('category items in this workbook')

#piechart
fig = plt.figure(figsize =(10, 7))
plt.pie(column_two_list, labels=column_one_list)
#plt.show()


#Create a dictionary from the two lists
#using nested for loop - naive method
dictone={}
for k in column_one_list:
    for v in column_two_list:
        dictone[k]=v
        column_two_list.remove(v)
        #if len(column_two_list):
            #break            
        #if bool(column_two_list):
            #break
        break
print('The dictionary is', dictone)     

